"""
created on 2020-02-12
changelog:
2021-05-14
-added error exception handling when there is issue with modifying excel file
-solved bug with not showing own ID when submitting ticket.

"""
import sys
import time
import xlrd
import openpyxl
import PySimpleGUI as sg
from selenium import webdriver
from openpyxl.styles import PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# import request
import json

# layout设计布局
# sg.theme('Dark Blue 3')#设计主题颜色
sg.theme('DarkAmber')
layout = [
    [sg.Text('Hi,please get your Ticket excel firstly:')],
    [sg.Input(), sg.FileBrowse()],  # values[0]
    [sg.Text('Please input your Start-Row:'), sg.Input(2)],  # values[1]
    [sg.Text('Please input your End-Row:'), sg.Input(3)],  # values[2]
    [sg.Text('Please input your work country:'),
     sg.Input("Germany")],  # values[3]
    [sg.Text('And your work city:'), sg.Input("Dusseldorf")],  # values[4]

    [sg.Button('Submit'), sg.Button('About'), sg.Text('  '),
     sg.Text('    Version 2.2 Created by Accenture B-team')]
]


# global element
def openTicketExcel(excel_Flie, start_Row, end_Row):
    workbook = xlrd.open_workbook(excel_Flie)  # 通过xlrd包打开提单表格，
    ticketSheet = workbook.sheet_by_name('Ticket Log')  # 打开提单页面
    ticketTable = []  # 使用二维数组存储Ticket表格
    # int(input("Please input your Start-Row:\n")) - 1  # 请输入本次要提单的开始行
    startRow = start_Row - 1
    # int(input("Please input your End-Row:\n"))  # 请输入本次要提单的结束行
    endRow = end_Row
    for row in range(startRow, endRow):
        Temp = ticketSheet.row_values(row)
        Time_Excel = Temp[6] + 1
        Time_Excel = (Time_Excel - 19 - 70 * 365) * \
            86400 - 8 * 3600  # 将excel时间序列值转化为时间戳
        Time_Local = time.localtime(Time_Excel)  # 将时间戳转化为localtime
        Temp[6] = time.strftime("%Y-%m-%d", Time_Local)  # 将localtime转换成提单日期
        ticketTable.append(Temp)  # 时间格式改好，记录入二维数据表
    workbook.release_resources()
    return ticketTable


def ticketSubmit(startRow, endRow, country, city, excel_Flie):
    for row in range(startRow, endRow + 1):
        sg.OneLineProgressMeter(
            'The progress of task execution',
            row - startRow + 1,
            endRow - startRow + 1,
            'key',
            'Optional message',
            orientation='h')
        try:
            element = WebDriverWait(
                driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="_tmL1"]/a')))
        finally:
            driver.find_element_by_xpath('//*[@id="_tmL1"]/a').click()
        try:
            element = WebDriverWait(driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, '//*[@id="applicationForm"]/div[1]/div[1]/div/input[2]')))
        finally:
            driver.find_element_by_xpath(
                '//*[@id="applicationForm"]/div[1]/div[1]/div/input[2]').click()
        # driver.find_element_by_id('userId').clear()  # 按照路径填入相应数据
        # driver.find_element_by_id('userId').send_keys(ticketTable[row - startRow][4])
        time.sleep(2)
        driver.find_element_by_id('userId').clear()
        time.sleep(1)
        driver.find_element_by_id('userId').send_keys(
            ticketTable[row - startRow][4])

        time.sleep(1)
        driver.find_element_by_id('country').send_keys(country)
        time.sleep(1)
        driver.find_element_by_id('city').send_keys(city)
        time.sleep(2)
        # //*[@id="applicationForm"]/div[1]/div[1]/ul/li[4]/span/a
        driver.find_element_by_xpath(
            '//*[@id="applicationForm"]/div[1]/div[1]/ul/li[4]/span/a').click()
        time.sleep(0.8)
        driver.find_element_by_xpath(
            '//*[@id="locationInfor_ddl"]/div[2]/ul/li[1]').click()
        # time.sleep(1)
        phone = driver.find_element_by_name('userPhone').get_attribute('value')
        if phone == '':
            driver.find_element_by_name('userPhone').send_keys('1')
        rome = driver.find_element_by_id('room').get_attribute('value')
        if rome == '':
            driver.find_element_by_id('room').send_keys('1')
        ticketTable_temp = list(ticketTable)
        for i in range(1, 13, 2):  # 向数组每个元素之后插入两个空格并写入文本框
            ticketTable_temp[row - startRow].insert(i, '  ')
        driver.find_element_by_name('faultDescription').send_keys(
            ticketTable_temp[row - startRow])
        time.sleep(1)
        driver.find_element_by_xpath(
            '//*[@id="btnSubmit"]/label').click()  # 点击提交按钮
        time.sleep(2)
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, Substring1 + str(row - startRow + 1) + Substring2))
                # 点击成功提交按钮
                # EC.presence_of_element_located((By.XPATH, '//*[@id="jalor_dialog1"]/div[3]/span'))#点击成功提交按钮
                # EC.text_to_be_present_in_element_value((By.CSS_SELECTOR,'jalor_dialog2'), u'OK')
            )
        finally:
            driver.find_element_by_xpath(
                Substring1 + str(row - startRow + 1) + Substring2).click()
            for col_index in range(1, 8):  # 在提单表格中进行进行标记该单
                try:
                    Get_mark = openpyxl.load_workbook(excel_Flie)
                    # Name_list = Get_mark.sheetnames  # print(Name_list)
                    fill = PatternFill("solid", fgColor='32cd32')
                    MarkTable = Get_mark['Ticket Log']
                    for col_num in range(1, 8):
                        MarkTable.cell(row=row, column=col_num).fill = fill
                    Get_mark.save(excel_Flie)
                except BaseException:
                    break

    driver.quit()


# create the window创建视窗
window = sg.Window('Automated ticketing system', layout)

if __name__ == "__main__":
    # 完成提单的准备工作：得到提单的EXCEL表格，输入起始行和结束行
    while True:

        event, values = window.Read()  # read the window读取视窗
        fname = values[0]
        start_Row = int(values[1])
        end_Row = int(values[2])
        country = str(values[3])
        city = str(values[4])

        if event == 'About':
            sg.Popup(
                "This tool is Created by Accenture B-Team member",
                "Changelog 6.4.2021",
                "Improved interface and solved some known bug",
                "If found issue, report bug to lee soon yik lwx705765",
                "Please note that the tool is developed by personal use, not meant to constantly maintain and improve",
                title="About")

        if event == 'Submit':
            if not fname:
                sg.Popup("No filename supplied")
                raise SystemExit("Cancelling: no filename supplied")
                break
            ticketTable = openTicketExcel(fname, start_Row, end_Row)
            # print(ticketTable[0])
            # 进行提单
            fill = PatternFill("solid", fgColor='32cd32')
            # Step1:打开网页并进入“IT onsite service下面的Others apply”
            driver = webdriver.Chrome('C:/Users/chromedriver.exe')
            driver.fullscreen_window()
            driver.get(
                'http://w3.huawei.com/btit/ebtittools/ios/#!har/home.html')
            # 开始进行循环提单:
            # 由于每次提单后确认按钮的XPATH会发生改变，所以需要对XPATH进行处理
            Substring1 = '//*[@id="jalor_dialog'
            Substring2 = '"]/div[3]/span'
            ticketSubmit(start_Row, end_Row, country, city, fname)
            break
